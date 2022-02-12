import pdb
from glob import glob, iglob
from pathlib import Path
from typing import Generator

import openpyxl
import os

from openpyxl.cell import Cell

from api import funcUtil


def scan_excels(base_dir):
    if not os.path.isdir(base_dir): raise NotADirectoryError("所传入参数不是合法路径")
    abp = Path(base_dir).absolute()
    afx = glob(f'{abp}/[Tt]est*.xlsx')
    if len(afx) < 1: raise FileNotFoundError("未找到所匹配的excel文件")
    return afx


def read_workbook(excel) -> dict:
    if not has_file(excel): raise FileNotFoundError(excel)

    wbs = openpyxl.load_workbook(excel)
    di = {}
    sheets = wbs.worksheets
    for sheet in sheets:
        sheet_data = []
        for val in sheet.values:
            sheet_data.append(val)
        di[sheet.title] = sheet_data
    wbs.close()
    yield {Path(excel).name: di}


def has_file(excel):
    return os.path.isfile(excel)


def write_workbook(excel, tuple_list, workbook='Sheet1'):
    wb = None
    if has_file(excel):
        wb = openpyxl.load_workbook(excel)
    else:
        wb = openpyxl.Workbook()


def __load_case_from_xldict(d: dict):
    '''
    :param d: excel表格中读取出来如下：
      g = {'Sheet1': [('说明', '步骤ID', '步骤名称', '关键字', '参数'), ('1.本说明列，必须在第一列', -1, '测试登录', None, None),
                    ('2.步骤ID主要用于分割每个测试用例，步骤ID为-1的，用于给出每个测试用例的基本信息，两个-1之间的即为本测试用例的操作步骤', 1, '跳转网页', 'get', None),
                    ('3.如果一个操作步骤有多个参数，从参数列开始，每个参数占据一个单元格，可以填写多列', 2, '输入', 'input', None)],
         'Sheet2': [(None, None, None), ('aa', 'bb', 'cc')]}
    :return:

        封装成如下结构
        c = {'module_name':
                 [{'casename1': 'name',
                   'case_no':1,
                   'steps':
                       [
                           ['step_name1', 'step_key', '*args'],
                           ['step_name2', 'step_key', '*args'],
                       ]
                   },
                  {'casename2': 'name',
                   'case_no':1,
                   'steps':
                       [
                           ['step_name1', 'step_key', '*args'],
                           ['step_name2', 'step_key', '*args'],
                       ]
                   },
                  ]
             }

    '''
    xl_name = list(d.keys())[0]
    xl_data = list(d.values())[0]
    all_case = {}
    for sheetname, tuple_list in xl_data.items():
        modulename = sheetname
        csls = {}
        case_no = 0

        for tup in tuple_list:
            if not isinstance(tup[1], int): continue
            if tup[1] == -1:
                case_no += 1
                test_case = {'xl_name': xl_name, 'sheetname': sheetname, 'case_no': case_no, 'case_name': tup[2],
                             'steps': []}
            else:
                test_case['steps'].append([_ for _ in tup[1:] if _])

            csls[case_no] = test_case

        all_case[modulename] = csls
    return {xl_name: all_case}


def test_suite_by_excel(base_dir) -> Generator:
    xlls = (scan_excels(base_dir))
    for xl in xlls:
        xl_dict_gen = read_workbook(xl)
        for di in xl_dict_gen:
            ca = __load_case_from_xldict(di)
            yield ca


def dec(ls):
    newls = []
    newdi = {}
    count = 0
    for item in ls:
        if item[0] == -1:
            newdi['no'] = count + 1
            newdi['step'] = []
            count = count + 1
        else:
            if count == newdi['no']:
                print(count)
                print(newdi)
                # if 'step' in newdi.keys():
                newdi['step'].append(item[1])
            # else:
            #     print(newdi)
    print(newdi)
    return newls


if __name__ == '__main__':
    # print(read_workbook(ls[1]))
    td = r"C:\Users\Uneven\Desktop\codes\python\seleniumWeb\excel"
    # has_file(td)
    # write_workbook(td,None)

    g = {'Sheet1': [('说明', '步骤ID', '步骤名称', '关键字', '参数'), ('1.本说明列，必须在第一列', -1, '测试登录', None, None),
                    ('2.步骤ID主要用于分割每个测试用例，步骤ID为-1的，用于给出每个测试用例的基本信息，两个-1之间的即为本测试用例的操作步骤', 1, '跳转网页', 'get', None),
                    ('3.如果一个操作步骤有多个参数，从参数列开始，每个参数占据一个单元格，可以填写多列', 2, '输入', 'input', None)],
         'Sheet2': [(None, None, None), ('aa', 'bb', 'cc')]}

    # for i in test_suite_by_excel(td):
    #     print(i)

    h = {'Sheet1': [{'case_no': 1, 'case_name': '测试登录', 'steps': [[1, '跳转网页', 'get'], [2, '输入', 'input']]},
                    {'case_no': 1, 'case_name': '测试登录', 'steps': [[1, '跳转网页', 'get'], [2, '输入', 'input']]},
                    {'case_no': 1, 'case_name': '测试登录', 'steps': [[1, '跳转网页', 'get'], [2, '输入', 'input']]}],
         'Sheet2': []}

    g = {'Test_ddt - 副本.xlsx': {'Sheet1': {}}}
    p = {'test_ddt.xlsx': {
        'Sheet1': {1: {'case_no': 1, 'case_name': '测试登录', 'steps': [[1, '跳转网页', 'get'], [2, '输入', 'input']]}},
        'Sheet2': {}}}
    ls = [
        [-1, 'a组'],
        [1, 'a-1'],
        [2, 'a-2'],
        [3, 'a-3'],
        [-1, 'b组'],
        [1, 'b-1'],
        [2, 'b-2'],
    ]
    d = {
        'no': 1,
        'step': [
            ['1', 1],
            ['2', 2]
        ]
    }
    print(dec(ls))
